import json
import random
import threading
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.db import models
from django.utils import timezone
from datetime import timedelta
from django.core.mail import send_mail
from django.conf import settings
from .models import UserAccount, RegistrationOTP, Attendance, Notice, Routine, RoutineFile, Assignment, AssignmentSubmission, ClassRepresentative, SemesterBook, Message, Complaint, Quiz, QuizQuestion, QuizSubmission

from django.core.mail import EmailMultiAlternatives

def send_html_otp_email(subject, title, otp_code, recipient_list):
    """Send OTP email in background thread for instant API response."""
    def _send():
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <body style="font-family: Arial, sans-serif; background: #f5f5f5; padding: 20px;">
          <div style="max-width: 480px; margin: auto; background: #fff; border-radius: 10px; padding: 30px; border: 1px solid #ddd;">
            <h2 style="color: #4f46e5; text-align: center;">🎓 College Management System</h2>
            <p style="color: #333; font-size: 15px;">{title}</p>
            <p style="color: #555; font-size: 14px;">Your verification code is:</p>
            <div style="text-align: center; margin: 25px 0;">
              <span style="font-size: 36px; font-weight: bold; letter-spacing: 8px; color: #4f46e5; font-family: monospace;">{otp_code}</span>
            </div>
            <p style="color: #888; font-size: 12px;">This code is valid for <strong>5 minutes</strong>. Do not share it with anyone.</p>
            <hr style="border: none; border-top: 1px solid #eee; margin: 20px 0;">
            <p style="color: #aaa; font-size: 11px; text-align: center;">College Management System &copy; 2025</p>
          </div>
        </body>
        </html>
        """
        text_content = f"{title}\n\nYour OTP code is: {otp_code}\n\nValid for 5 minutes. Do not share this code.\n\n- College Management System"
        try:
            msg = EmailMultiAlternatives(
                subject=subject,
                body=text_content,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=recipient_list,
            )
            msg.attach_alternative(html_content, "text/html")
            msg.send(fail_silently=False)
            print(f"[EMAIL OK] Sent to {recipient_list}")
        except Exception as e:
            print(f"[EMAIL FAIL] {str(e)}")

    threading.Thread(target=_send, daemon=True).start()



@csrf_exempt
def send_registration_otp(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email = data.get('email')
            role = data.get('role')
            pin = data.get('pin')
            
            # Admin PIN validation for Teachers
            if role == 'teacher':
                if pin != '730523':
                    return JsonResponse({'status': 'error', 'message': 'Invalid Admin Master PIN!'}, status=403)

            # Check if email is already in use
            if UserAccount.objects.filter(email=email).exists():
                return JsonResponse({'status': 'error', 'message': 'Email already registered!'}, status=400)
            
            otp_code = str(random.randint(100000, 999999))
            
            RegistrationOTP.objects.update_or_create(
                email=email,
                defaults={
                    'otp': otp_code,
                    'otp_expiry': timezone.now() + timedelta(minutes=5)
                }
            )
            
            # Print OTP to console as fallback (always visible in server terminal)
            print(f"")
            print(f"{'='*50}")
            print(f"  [REGISTRATION OTP] Email: {email}")
            print(f"  [REGISTRATION OTP] Code : {otp_code}")
            print(f"{'='*50}")
            print(f"")
            
            # Send Email
            subject = f"{otp_code} is your College MS verification code"
            title = 'Account Registration'
            send_html_otp_email(subject, title, otp_code, [email])
            
            return JsonResponse({'status': 'success', 'message': 'OTP sent to your email!'})
        except Exception as e:
            print(f"[ERROR SENDING OTP]: {str(e)}")
            return JsonResponse({'status': 'error', 'message': f"Failed to send email: {str(e)}"}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@csrf_exempt
def register(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email = data.get('email')
            otp_code = data.get('registration_otp')
            
            # Verify OTP
            reg_otp = RegistrationOTP.objects.filter(email=email).first()
            if not reg_otp:
                 return JsonResponse({'status': 'error', 'message': 'Please request an OTP first!'}, status=400)
            
            if reg_otp.otp != otp_code:
                return JsonResponse({'status': 'error', 'message': 'Incorrect OTP!'}, status=400)
                
            if timezone.now() > reg_otp.otp_expiry:
                return JsonResponse({'status': 'error', 'message': 'OTP has expired!'}, status=400)

            # Check again if email exists (race condition)
            if UserAccount.objects.filter(email=email).exists():
                return JsonResponse({'status': 'error', 'message': 'Email already exists!'}, status=400)
            
            user = UserAccount(
                role=data.get('role'),
                first_name=data.get('first_name'),
                last_name=data.get('last_name'),
                mobile=data.get('mobile'),
                email=email,
                gender=data.get('gender'),
                password=data.get('password'),
                department=data.get('department'),
                roll=data.get('roll'),
                batch=data.get('batch'),
                session=data.get('session'),
                semester=data.get('semester')
            )
            user.save()
            
            # Clear Registration OTP after success
            reg_otp.delete()
            
            return JsonResponse({'status': 'success', 'message': 'Account created successfully!'})
        except Exception as e:
            print(f"[REGISTER ERROR]: {str(e)}")
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@csrf_exempt
def register_student_by_teacher(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email = data.get('email')
            
            # Check if email already exists
            if UserAccount.objects.filter(email=email).exists():
                return JsonResponse({'status': 'error', 'message': 'Email already exists!'}, status=400)
            
            user = UserAccount(
                role='student',
                first_name=data.get('first_name'),
                last_name=data.get('last_name'),
                mobile=data.get('mobile'),
                email=email,
                gender=data.get('gender'),
                password=data.get('password', 'student123'),
                department=data.get('department'),
                roll=data.get('roll'),
                batch=data.get('batch'),
                session=data.get('session'),
                semester=data.get('semester')
            )
            user.save()
            
            return JsonResponse({'status': 'success', 'message': 'Student registered successfully!'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@csrf_exempt
def login(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email_or_mobile = data.get('email_or_mobile')
            password = data.get('password')
            role = data.get('role')
            
            # Check if user exists by email or mobile
            user_query = UserAccount.objects.filter(role=role).filter(
                models.Q(email=email_or_mobile) | models.Q(mobile=email_or_mobile)
            )
            
            if not user_query.exists():
                return JsonResponse({'status': 'error', 'message': 'Email not found'}, status=404)
            
            u = user_query.first()
            if u.password != password:
                return JsonResponse({'status': 'error', 'message': 'Wrong password'}, status=401)
            
            # Direct login for both roles (no OTP on login, only during registration)
            return JsonResponse({
                'status': 'success', 
                'message': 'Login successful!',
                'data': {
                    'first_name': u.first_name,
                    'last_name': u.last_name,
                    'email': u.email,
                    'role': u.role,
                    'roll': getattr(u, 'roll', '') or '',
                    'mobile': getattr(u, 'mobile', '') or '',
                    'department': getattr(u, 'department', '') or '',
                    'semester': getattr(u, 'semester', '') or '',
                    'shift': getattr(u, 'shift', '') or '',
                    'profile_picture': u.profile_picture.url if u.profile_picture else None
                }
            })
            
            # Student login - direct (no OTP)
            return JsonResponse({
                'status': 'success', 
                'message': 'Login successful!',
                'data': {
                    'first_name': u.first_name,
                    'last_name': u.last_name,
                    'email': u.email,
                    'role': u.role,
                    'profile_picture': u.profile_picture.url if u.profile_picture else None
                }
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@csrf_exempt
def verify_login_otp(request):
    """Verify OTP for teacher login and return user data on success."""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email = data.get('email')
            otp_code = data.get('otp')
            
            user = UserAccount.objects.filter(email=email, role='teacher').first()
            if not user:
                return JsonResponse({'status': 'error', 'message': 'Invalid user!'}, status=404)
            
            if user.otp != otp_code:
                return JsonResponse({'status': 'error', 'message': 'Incorrect OTP!'}, status=400)
            
            if timezone.now() > user.otp_expiry:
                return JsonResponse({'status': 'error', 'message': 'OTP has expired!'}, status=400)
            
            # Clear OTP after successful verification
            user.otp = None
            user.otp_expiry = None
            user.save()
            
            return JsonResponse({
                'status': 'success',
                'message': 'Login successful!',
                'data': {
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'email': user.email,
                    'role': user.role,
                    'profile_picture': user.profile_picture.url if user.profile_picture else None
                }
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@csrf_exempt
def send_otp(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email = data.get('email')
            
            user = UserAccount.objects.filter(email=email).first()
            if not user:
                return JsonResponse({'status': 'error', 'message': 'No account found with this email!'}, status=404)
            
            otp_code = str(random.randint(100000, 999999))
            user.otp = otp_code
            user.otp_expiry = timezone.now() + timedelta(minutes=5)
            user.save()
            
            # Send HTML Email Synchronously
            subject = f"{otp_code} is your College MS password reset code"
            title = 'Password Reset'
            send_html_otp_email(subject, title, otp_code, [email])
            
            return JsonResponse({'status': 'success', 'message': 'OTP sent to your email!'})
        except Exception as e:
            print(f"[ERROR SENDING RESET OTP]: {str(e)}")
            return JsonResponse({'status': 'error', 'message': f"Failed to send email: {str(e)}"}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@csrf_exempt
def verify_otp(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email = data.get('email')
            otp_code = data.get('otp')
            
            user = UserAccount.objects.filter(email=email).first()
            if not user:
                return JsonResponse({'status': 'error', 'message': 'Invalid User!'}, status=404)
            
            if user.otp != otp_code:
                return JsonResponse({'status': 'error', 'message': 'Incorrect OTP!'}, status=400)
            
            if timezone.now() > user.otp_expiry:
                return JsonResponse({'status': 'error', 'message': 'OTP has expired!'}, status=400)
            
            return JsonResponse({'status': 'success', 'message': 'OTP verified successfully!'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@csrf_exempt
def reset_password(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email = data.get('email')
            otp_code = data.get('otp')
            new_password = data.get('new_password')
            
            user = UserAccount.objects.filter(email=email).first()
            if not user:
                return JsonResponse({'status': 'error', 'message': 'Invalid User!'}, status=404)
                
            if user.otp != otp_code or timezone.now() > user.otp_expiry:
                return JsonResponse({'status': 'error', 'message': 'OTP invalid or expired!'}, status=400)
                
            user.password = new_password
            user.otp = None         # clear OTP
            user.otp_expiry = None
            user.save()
            
            return JsonResponse({
                'status': 'success', 
                'message': 'Password reset successful!',
                'data': {
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'role': user.role
                }
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@csrf_exempt
def get_students_for_attendance(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            department = data.get('department')
            semester = data.get('semester')
            session = data.get('session')
            date_str = data.get('date')
            subject_name = data.get('subject')
            subject_code = data.get('subject_code', '')
            final_subject = f"{subject_name} ({subject_code})" if subject_code else subject_name
            
            query = UserAccount.objects.filter(role='student')

            if department and department != 'All':
                dep_keyword = department.split()[0]
                query = query.filter(department__icontains=dep_keyword)

            if semester and semester != 'All':
                sem_first_word = semester.split()[0] # e.g. "5th", "7th", "1st"
                query = query.filter(semester__icontains=sem_first_word)

            if session and session != 'All':
                # Try filtering by session first
                session_query = query.filter(models.Q(session__icontains=session) | models.Q(session__icontains=session.split('-')[0]))
                if session_query.exists():
                    query = session_query

            student_list = []
            for s in query.order_by('roll', 'first_name'):
                # Check for existing attendance
                existing_att = Attendance.objects.filter(student=s, date=date_str, subject=final_subject).first()
                status = existing_att.status if existing_att else None
                
                student_list.append({
                    'id': s.id, 
                    'first_name': s.first_name, 
                    'last_name': s.last_name, 
                    'roll': s.roll or '--',
                    'status': status
                })
                
            return JsonResponse({'status': 'success', 'data': student_list})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@csrf_exempt
def save_attendance(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            attendance_list = data.get('attendance')
            subject_name = data.get('subject')
            subject_code = data.get('subject_code', '')
            final_subject = f"{subject_name} ({subject_code})" if subject_code else subject_name
            
            semester = data.get('semester')
            session = data.get('session')
            date_str = data.get('date')
            for item in attendance_list:
                student = UserAccount.objects.get(id=item['student_id'])
                Attendance.objects.update_or_create(
                    student=student, date=date_str, subject=final_subject,
                    defaults={'status': item['status'], 'batch': semester, 'session': session}
                )
            return JsonResponse({'status': 'success', 'message': 'Attendance saved successfully!'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@csrf_exempt
def get_teacher_dashboard_data(request):
    if request.method == 'GET':
        try:
            total_students = UserAccount.objects.filter(role='student').count()
            today = timezone.now().date()
            attendance_today = Attendance.objects.filter(date=today, status__in=['P', 'L']).count()
            notices = Notice.objects.all().order_by('-date_posted')[:5]
            notice_list = [{
                'id': n.id, 
                'title': n.title, 
                'content': n.content, 
                'date': n.date_posted.strftime('%d %b, %Y'),
                'attachment': n.attachment.url if n.attachment else None
            } for n in notices]
            day_name = timezone.now().strftime('%A')
            routines = Routine.objects.filter(day=day_name).order_by('time_slot')
            routine_list = [{'subject': r.subject, 'time': r.time_slot, 'room': r.room, 'batch': r.batch} for r in routines]
            return JsonResponse({
                'status': 'success',
                'data': {
                    'total_students': total_students,
                    'attendance_today': attendance_today,
                    'classes_today': routines.count(),
                    'notices': notice_list,
                    'routines': routine_list
                }
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@csrf_exempt
def get_profile(request):
    if request.method in ['POST', 'GET']:
        try:
            if request.method == 'POST':
                try:
                    data = json.loads(request.body)
                    email = data.get('email')
                    roll = data.get('roll')
                except Exception:
                    email = request.POST.get('email')
                    roll = request.POST.get('roll')
            else:
                email = request.GET.get('email')
                roll = request.GET.get('roll')

            user = None
            if email:
                user = UserAccount.objects.filter(email=email).first()
            if not user and roll:
                user = UserAccount.objects.filter(roll=roll).first()

            if not user:
                return JsonResponse({'status': 'error', 'message': 'User not found'}, status=404)

            return JsonResponse({
                'status': 'success',
                'data': {
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'email': user.email,
                    'mobile': user.mobile or '',
                    'roll': user.roll or '',
                    'department': user.department or '',
                    'semester': user.semester or '',
                    'shift': getattr(user, 'shift', '') or '',
                    'room_number': getattr(user, 'room_number', '') or '',
                    'qualification': getattr(user, 'qualification', '') or '',
                    'specialized_subjects': getattr(user, 'specialized_subjects', '') or '',
                    'assigned_classes': getattr(user, 'assigned_classes', '') or '',
                    'profile_picture': user.profile_picture.url if user.profile_picture else None
                }
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@csrf_exempt
def update_profile(request):
    if request.method == 'POST':
        try:
            # Multi-part form data or JSON
            current_email = request.POST.get('current_email') or request.POST.get('email')
            roll = request.POST.get('roll')

            user = None
            if current_email:
                user = UserAccount.objects.filter(email=current_email).first()
            if not user and roll:
                user = UserAccount.objects.filter(roll=roll).first()

            if not user:
                return JsonResponse({'status': 'error', 'message': 'User account not found.'}, status=404)

            first_name = request.POST.get('first_name')
            if first_name is not None and first_name.strip():
                user.first_name = first_name.strip()

            last_name = request.POST.get('last_name')
            if last_name is not None and last_name.strip():
                user.last_name = last_name.strip()

            new_email = request.POST.get('new_email') or request.POST.get('email')
            if new_email and new_email.strip() and new_email.strip() != user.email:
                if not UserAccount.objects.filter(email=new_email.strip()).exclude(id=user.id).exists():
                    user.email = new_email.strip()

            mobile = request.POST.get('mobile')
            if mobile is not None and mobile.strip():
                user.mobile = mobile.strip()

            if 'profile_picture' in request.FILES:
                user.profile_picture = request.FILES['profile_picture']

            user.save()

            pic_url = user.profile_picture.url if user.profile_picture else None
            return JsonResponse({
                'status': 'success',
                'message': 'Profile updated successfully!',
                'profile_picture': pic_url,
                'data': {
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'email': user.email,
                    'mobile': user.mobile or '',
                    'roll': user.roll or '',
                    'department': user.department or '',
                    'semester': user.semester or '',
                    'profile_picture': pic_url
                }
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@csrf_exempt
def change_password(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email = data.get('email')
            current_password = data.get('current_password')
            new_password = data.get('new_password')
            
            user = UserAccount.objects.get(email=email)
            if user.password != current_password:
                return JsonResponse({'status': 'error', 'message': 'Current password is incorrect!'}, status=400)
            
            user.password = new_password
            user.save()
            return JsonResponse({'status': 'success', 'message': 'Password changed successfully!'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@csrf_exempt
def get_teachers(request):
    if request.method == 'GET':
        try:
            teachers = UserAccount.objects.filter(role='teacher')
            teacher_list = []
            for t in teachers:
                # Fetch distinct subjects this teacher teaches from the routine
                subjects = list(Routine.objects.filter(teacher=t).values_list('subject', flat=True).distinct())
                teacher_list.append({
                    'name': f"{t.first_name} {t.last_name}",
                    'department': t.department,
                    'email': t.email,
                    'mobile': t.mobile,
                    'designation': t.designation or 'Faculty',
                    'room_number': t.room_number,
                    'qualification': t.qualification,
                    'specialized_subjects': t.specialized_subjects,
                    'assigned_classes': t.assigned_classes,
                    'subjects': subjects,
                    'profile_picture': t.profile_picture.url if t.profile_picture else None
                })
            return JsonResponse({'status': 'success', 'data': teacher_list})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@csrf_exempt
def delete_student(request, student_id):
    if request.method == 'DELETE':
        try:
            student = UserAccount.objects.get(id=student_id, role='student')
            student.delete()
            return JsonResponse({'status': 'success', 'message': 'Student deleted successfully!'})
        except UserAccount.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Student not found'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@csrf_exempt
def get_students(request):
    if request.method in ['GET', 'POST']:
        try:
            department = None
            session = None
            semester = None
            if request.method == 'POST':
                try:
                    data = json.loads(request.body)
                    department = data.get('department')
                    session = data.get('session')
                    semester = data.get('semester')
                except Exception:
                    pass
            else:
                department = request.GET.get('department')
                session = request.GET.get('session')
                semester = request.GET.get('semester')
            
            query = UserAccount.objects.filter(role='student')
            if department and department != 'All': 
                dep_keyword = department.split()[0]
                query = query.filter(department__icontains=dep_keyword)
            if session: query = query.filter(session=session)
            
            if semester:
                from django.db.models import Q
                # Map common semester names to ensure matches
                if "1st" in semester or "1" in semester: sem_q = Q(semester="1st Semester") | Q(semester="Semester 1")
                elif "2nd" in semester or "2" in semester: sem_q = Q(semester="2nd Semester") | Q(semester="Semester 2")
                elif "3rd" in semester or "3" in semester: sem_q = Q(semester="3rd Semester") | Q(semester="Semester 3")
                elif "4th" in semester or "4" in semester: sem_q = Q(semester="4th Semester") | Q(semester="Semester 4")
                elif "5th" in semester or "5" in semester: sem_q = Q(semester="5th Semester") | Q(semester="Semester 5")
                elif "6th" in semester or "6" in semester: sem_q = Q(semester="6th Semester") | Q(semester="Semester 6")
                elif "7th" in semester or "7" in semester: sem_q = Q(semester="7th Semester") | Q(semester="Semester 7")
                elif "8th" in semester or "8" in semester: sem_q = Q(semester="8th Semester") | Q(semester="Semester 8")
                else: sem_q = Q(semester=semester)
                query = query.filter(sem_q)
            
            student_list = [{
                'id': s.id,
                'roll': s.roll,
                'name': f"{s.first_name} {s.last_name}",
                'department': s.department,
                'session': s.session,
                'semester': s.semester,
                'profile_picture': s.profile_picture.url if s.profile_picture else None
            } for s in query]
            return JsonResponse({'status': 'success', 'data': student_list})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@csrf_exempt
def search_users(request):
    query = request.GET.get('q', '')
    if not query:
        return JsonResponse({'status': 'error', 'message': 'Empty search query'}, status=400)
    
    try:
        # Search by ID, Roll or Name
        search_filter = models.Q(roll__icontains=query) | models.Q(first_name__icontains=query) | models.Q(last_name__icontains=query)
        
        if query.isdigit():
            search_filter |= models.Q(id=query)
            
        results = UserAccount.objects.filter(search_filter)
        
        data = [{
            'id': r.id,
            'name': f"{r.first_name} {r.last_name}",
            'role': r.role,
            'roll': r.roll,
            'email': r.email,
            'department': r.department,
            'session': r.session,
            'semester': r.semester,
            'profile_picture': r.profile_picture.url if r.profile_picture else None
        } for r in results]
        return JsonResponse({'status': 'success', 'data': data})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def get_student_stats(request):
    student_id = request.GET.get('student_id')
    if not student_id:
        return JsonResponse({'status': 'error', 'message': 'Missing student_id'}, status=400)
        
    try:
        student = UserAccount.objects.get(id=student_id, role='student')
        # Aggregate stats from Attendance model
        attendance_set = Attendance.objects.filter(student=student)
        total_classes = attendance_set.count()
        missed_classes = attendance_set.filter(status='A').count()
        present_late = attendance_set.filter(status__in=['P', 'L']).count()
        
        if total_classes == 0:
            rate = 0.0
            performance = "Normal"
        else:
            rate = (present_late / total_classes * 100)
            if rate >= 80:
                performance = "Verry Good"
            elif rate >= 60:
                performance = "Good"
            else:
                performance = "Bad"
            
        return JsonResponse({
            'status': 'success',
            'data': {
                'total_classes': total_classes,
                'attended_classes': present_late,
                'missed_classes': missed_classes,
                'attendance_rate': f"{rate:.1f}%",
                'performance': performance
            }
        })
    except UserAccount.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Student not found'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def get_monthly_report(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            department = data.get('department')
            semester = data.get('semester')
            session = data.get('session')
            subject_name = data.get('subject') # Optional filtering
            date_str = data.get('date')
            
            if not date_str:
                return JsonResponse({'status': 'error', 'message': 'Date is required to determine the month'}, status=400)
                
            year, month, _ = date_str.split('-')
            
            # Base query for students
            student_query = UserAccount.objects.filter(role='student', session=session, semester=semester)
            if department:
                dep_keyword = department.split()[0]
                student_query = student_query.filter(department__icontains=dep_keyword)
                
            report_data = []
            
            # For each student, calculate their stats for that month
            for s in student_query:
                att_query = Attendance.objects.filter(student=s, date__year=year, date__month=month)
                if subject_name:
                    att_query = att_query.filter(subject__istartswith=subject_name)
                    
                total = att_query.count()
                present = att_query.filter(status='P').count()
                late = att_query.filter(status='L').count()
                absent = att_query.filter(status='A').count()
                rate = ((present + late) / total * 100) if total > 0 else 0.0
                
                report_data.append({
                    'id': s.id,
                    'roll': s.roll,
                    'name': f"{s.first_name} {s.last_name}",
                    'total': total,
                    'present': present,
                    'late': late,
                    'absent': absent,
                    'rate': f"{rate:.1f}%"
                })
                
            return JsonResponse({
                'status': 'success', 
                'month': f"{year}-{month}",
                'data': report_data
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@csrf_exempt
def get_all_notices(request):
    if request.method == 'GET':
        try:
            dept = request.GET.get('department')
            sem = request.GET.get('semester')

            query = Notice.objects.all()

            if dept and dept != 'All':
                query = query.filter(models.Q(target_department='All') | models.Q(target_department__icontains=dept.split()[0]))
            if sem and sem != 'All':
                query = query.filter(models.Q(target_semester='All') | models.Q(target_semester__icontains=sem.split()[0]))

            notices = query.order_by('-date_posted')
            notice_list = [{
                'id': n.id,
                'title': n.title,
                'content': n.content,
                'target_department': getattr(n, 'target_department', 'All'),
                'target_semester': getattr(n, 'target_semester', 'All'),
                'date': n.date_posted.strftime('%d %b, %Y'),
                'posted_by': f"{n.posted_by.first_name} {n.posted_by.last_name}" if n.posted_by else "Admin",
                'attachment': n.attachment.url if n.attachment else None
            } for n in notices]
            return JsonResponse({'status': 'success', 'data': notice_list})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@csrf_exempt
def create_notice(request):
    if request.method == 'POST':
        try:
            email = request.POST.get('email')
            title = request.POST.get('title')
            content = request.POST.get('content')
            target_dept = request.POST.get('target_department', 'All')
            target_sem = request.POST.get('target_semester', 'All')
            attachment = request.FILES.get('attachment')
            
            if not title or not content:
                return JsonResponse({'status': 'error', 'message': 'Title and content are required'}, status=400)
                
            teacher = UserAccount.objects.filter(email=email, role='teacher').first()
            if not teacher:
                return JsonResponse({'status': 'error', 'message': 'Unauthorized to post notice'}, status=403)
                
            n = Notice(
                title=title, 
                content=content, 
                target_department=target_dept,
                target_semester=target_sem,
                posted_by=teacher, 
                attachment=attachment
            )
            n.save()
            return JsonResponse({'status': 'success', 'message': 'Notice created successfully!'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@csrf_exempt
def delete_notice(request, notice_id):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email = data.get('email')
            
            # Verify if the user is a teacher
            teacher = UserAccount.objects.filter(email=email, role='teacher').first()
            if not teacher:
                return JsonResponse({'status': 'error', 'message': 'Unauthorized. Only teachers can delete notices.'}, status=403)
            
            notice = Notice.objects.get(id=notice_id)
            notice.delete()
            return JsonResponse({'status': 'success', 'message': 'Notice deleted successfully!'})
        except Notice.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Notice not found.'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


@csrf_exempt
def edit_notice(request, notice_id):
    """Teacher edits an existing notice."""
    if request.method == 'POST':
        try:
            email = request.POST.get('email')
            teacher = UserAccount.objects.filter(email=email, role='teacher').first()
            if not teacher:
                teacher = UserAccount.objects.filter(email=email).first()
            if not teacher:
                teacher = UserAccount.objects.filter(role='teacher').first()
            if not teacher:
                return JsonResponse({'status': 'error', 'message': 'Unauthorized teacher'}, status=403)

            notice = Notice.objects.get(id=notice_id)
            title = request.POST.get('title', notice.title)
            content = request.POST.get('content', notice.content)
            target_dept = request.POST.get('target_department', notice.target_department)
            target_sem = request.POST.get('target_semester', notice.target_semester)
            attachment = request.FILES.get('attachment')

            notice.title = title
            notice.content = content
            notice.target_department = target_dept
            notice.target_semester = target_sem
            if attachment:
                notice.attachment = attachment
            notice.save()
            return JsonResponse({'status': 'success', 'message': 'Notice updated successfully!'})
        except Notice.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Notice not found.'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

# ==============================================================================
# ROUTINE MANAGEMENT SYSTEM API
# ==============================================================================

DEMO_ROUTINE_SLOTS = [
    # Sunday
    { 'day': 'Sunday', 'start_time': '08:45', 'end_time': '09:30', 'time_slot': '08:45-09:30', 'subject_code': '28572', 'subject': 'Network', 'teacher_initials': 'RH', 'teacher_name': 'Engr. Ripon Hossain', 'room': 'Com-304' },
    { 'day': 'Sunday', 'start_time': '09:30', 'end_time': '10:15', 'time_slot': '09:30-10:15', 'subject_code': '25853', 'subject': 'Innovation', 'teacher_initials': 'ShK', 'teacher_name': 'Shahin Khan', 'room': 'Com-304' },
    { 'day': 'Sunday', 'start_time': '10:15', 'end_time': '11:00', 'time_slot': '10:15-11:00', 'subject_code': '28573', 'subject': 'Cyber', 'teacher_initials': 'YA', 'teacher_name': 'Yeasin Arafat', 'room': 'Com-304' },
    { 'day': 'Sunday', 'start_time': '11:45', 'end_time': '01:15', 'time_slot': '11:45-01:15', 'subject_code': '28571', 'subject': 'Digital Marketing', 'teacher_initials': 'AL', 'teacher_name': 'Ali Hossain', 'room': 'Network - Lab' },

    # Monday
    { 'day': 'Monday', 'start_time': '08:00', 'end_time': '09:30', 'time_slot': '08:00-09:30', 'subject_code': '28575', 'subject': 'Multimedia', 'teacher_initials': 'RD', 'teacher_name': 'R S Ridoy', 'room': 'Com-304' },
    { 'day': 'Monday', 'start_time': '09:30', 'end_time': '11:00', 'time_slot': '09:30-11:00', 'subject_code': '28572', 'subject': 'Network', 'teacher_initials': 'RH', 'teacher_name': 'Engr. Ripon Hossain', 'room': 'Com-304' },
    { 'day': 'Monday', 'start_time': '11:45', 'end_time': '01:15', 'time_slot': '11:45-01:15', 'subject_code': '28576', 'subject': 'Project Work-2', 'teacher_initials': 'FH', 'teacher_name': 'Farhana Hossain', 'room': 'IOT - Lab' },

    # Tuesday
    { 'day': 'Tuesday', 'start_time': '08:00', 'end_time': '09:30', 'time_slot': '08:00-09:30', 'subject_code': '28575', 'subject': 'Multimedia', 'teacher_initials': 'RD', 'teacher_name': 'R S Ridoy', 'room': 'Hardware - Lab' },
    { 'day': 'Tuesday', 'start_time': '10:15', 'end_time': '11:00', 'time_slot': '10:15-11:00', 'subject_code': '28573', 'subject': 'Cyber', 'teacher_initials': 'YA', 'teacher_name': 'Yeasin Arafat', 'room': 'Com-304' },
    { 'day': 'Tuesday', 'start_time': '11:45', 'end_time': '01:15', 'time_slot': '11:45-01:15', 'subject_code': '28572', 'subject': 'Network', 'teacher_initials': 'RH', 'teacher_name': 'Engr. Ripon Hossain', 'room': 'Network-Lab' },

    # Wednesday
    { 'day': 'Wednesday', 'start_time': '08:00', 'end_time': '09:30', 'time_slot': '08:00-09:30', 'subject_code': '28574', 'subject': 'Apps Development', 'teacher_initials': 'SR', 'teacher_name': 'R S Ridoy Khan', 'room': 'Network-Lab' },
    { 'day': 'Wednesday', 'start_time': '10:15', 'end_time': '11:00', 'time_slot': '10:15-11:00', 'subject_code': '28571', 'subject': 'Digital-Marketing', 'teacher_initials': 'AL', 'teacher_name': 'Ali Hossain', 'room': 'Com-304' },
    { 'day': 'Wednesday', 'start_time': '11:45', 'end_time': '01:15', 'time_slot': '11:45-01:15', 'subject_code': '28573', 'subject': 'Cyber', 'teacher_initials': 'YA', 'teacher_name': 'Yeasin Arafat', 'room': 'Network-Lab' },

    # Thursday
    { 'day': 'Thursday', 'start_time': '08:00', 'end_time': '09:30', 'time_slot': '08:00-09:30', 'subject_code': '28576', 'subject': 'Project Work-2', 'teacher_initials': 'FH', 'teacher_name': 'Farhana Hossain', 'room': 'IOT - Lab' },
    { 'day': 'Thursday', 'start_time': '10:15', 'end_time': '11:00', 'time_slot': '10:15-11:00', 'subject_code': '28571', 'subject': 'Digital-Marketing', 'teacher_initials': 'AL', 'teacher_name': 'Ali Hossain', 'room': 'Com-304' },
    { 'day': 'Thursday', 'start_time': '11:00', 'end_time': '11:45', 'time_slot': '11:00-11:45', 'subject_code': '28574', 'subject': 'Apps Development', 'teacher_initials': 'SR', 'teacher_name': 'R S Ridoy Khan', 'room': 'Network-Lab' },
    { 'day': 'Thursday', 'start_time': '11:45', 'end_time': '12:30', 'time_slot': '11:45-12:30', 'subject_code': '25853', 'subject': 'Innovation', 'teacher_initials': 'ShK', 'teacher_name': 'Shahin Khan', 'room': 'Com-304' },
]

@csrf_exempt
def seed_demo_routine(request):
    """Seed DB with the 1st Shift Class Routine provided in the image if empty or forced."""
    try:
        force = request.GET.get('force') == 'true' or (request.body and json.loads(request.body).get('force'))
        if force:
            Routine.objects.filter(shift='1st Shift', department='Computer Science & Technology').delete()

        if not Routine.objects.filter(shift='1st Shift').exists() or force:
            for item in DEMO_ROUTINE_SLOTS:
                Routine.objects.create(
                    department='Computer Science & Technology',
                    semester='5th Semester',
                    shift='1st Shift',
                    section='A',
                    subject_code=item['subject_code'],
                    subject=item['subject'],
                    teacher_initials=item['teacher_initials'],
                    teacher_name=item['teacher_name'],
                    day=item['day'],
                    start_time=item['start_time'],
                    end_time=item['end_time'],
                    time_slot=item['time_slot'],
                    room=item['room'],
                    session='2023-24',
                    academic_year='2025-2026'
                )
            return JsonResponse({'status': 'success', 'message': f'Demo 1st Shift routine seeded successfully ({len(DEMO_ROUTINE_SLOTS)} slots)!'})
        return JsonResponse({'status': 'success', 'message': 'Routine data already exists.'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def get_routine(request):
    """Get routines with filtering and live class calculations (current class & next class)."""
    try:
        # Seed demo data automatically if DB is empty
        if not Routine.objects.exists():
            seed_demo_routine(request)

        department = request.GET.get('department')
        semester = request.GET.get('semester')
        shift = request.GET.get('shift')
        teacher_query = request.GET.get('teacher') # teacher email, initials, or name
        day_filter = request.GET.get('day')
        search_query = request.GET.get('search')

        query = Routine.objects.all()

        if department and department != 'All':
            query = query.filter(department__icontains=department.split()[0])
        if semester and semester != 'All':
            query = query.filter(semester__icontains=semester.split()[0])
        if shift and shift != 'All':
            query = query.filter(shift__icontains=shift.split()[0])
        if day_filter and day_filter != 'All':
            query = query.filter(day__iexact=day_filter)

        if teacher_query and teacher_query != 'All':
            query = query.filter(
                models.Q(teacher_initials__icontains=teacher_query) |
                models.Q(teacher_name__icontains=teacher_query) |
                models.Q(teacher__email__icontains=teacher_query) |
                models.Q(teacher__first_name__icontains=teacher_query) |
                models.Q(teacher__last_name__icontains=teacher_query)
            )

        if search_query:
            query = query.filter(
                models.Q(subject__icontains=search_query) |
                models.Q(subject_code__icontains=search_query) |
                models.Q(room__icontains=search_query) |
                models.Q(teacher_initials__icontains=search_query)
            )

        routines = list(query.order_by('day', 'start_time'))

        # Prepare serializable list
        routine_list = []
        for r in routines:
            routine_list.append({
                'id': r.id,
                'department': r.department,
                'semester': r.semester,
                'shift': r.shift,
                'section': r.section,
                'subject_code': r.subject_code or '',
                'subject': r.subject,
                'teacher_initials': r.teacher_initials or '',
                'teacher_name': r.teacher_name or (f"{r.teacher.first_name} {r.teacher.last_name}" if r.teacher else ''),
                'day': r.day,
                'start_time': r.start_time,
                'end_time': r.end_time,
                'time_slot': r.time_slot or f"{r.start_time}-{r.end_time}",
                'room': r.room,
                'session': r.session,
                'academic_year': r.academic_year
            })

        # Calculate Current Class & Next Class based on server time (or client passed time)
        client_day = request.GET.get('client_day') or timezone.now().strftime('%A')
        client_time = request.GET.get('client_time') or timezone.now().strftime('%H:%M')

        current_class = None
        next_class = None
        upcoming_alert = None  # 10-min advance warning

        today_routines = [r for r in routine_list if r['day'].lower() == client_day.lower()]
        today_routines.sort(key=lambda x: x['start_time'])

        # Parse client_time to minutes for comparison
        try:
            ct_parts = client_time.split(':')
            ct_minutes = int(ct_parts[0]) * 60 + int(ct_parts[1])
        except:
            ct_minutes = 0

        for r in today_routines:
            st = r['start_time']
            et = r['end_time']
            if st <= client_time <= et:
                current_class = r
            elif st > client_time and not next_class:
                next_class = r

        # Calculate upcoming_alert: if next class starts within 10 minutes
        if next_class and not current_class:
            try:
                nst_parts = next_class['start_time'].split(':')
                nst_minutes = int(nst_parts[0]) * 60 + int(nst_parts[1])
                diff = nst_minutes - ct_minutes
                if 0 < diff <= 10:
                    upcoming_alert = {
                        'class': next_class,
                        'minutes_left': diff,
                        'message': f"⚡ {next_class['subject']} class starts in {diff} minute{'s' if diff > 1 else ''}!"
                    }
            except:
                pass

        # Calculate minutes_until for next_class
        minutes_until_next = None
        if next_class:
            try:
                nst_parts = next_class['start_time'].split(':')
                nst_minutes = int(nst_parts[0]) * 60 + int(nst_parts[1])
                minutes_until_next = nst_minutes - ct_minutes
            except:
                pass

        return JsonResponse({
            'status': 'success',
            'data': routine_list,
            'current_class': current_class,
            'next_class': next_class,
            'upcoming_alert': upcoming_alert,
            'minutes_until_next': minutes_until_next,
            'current_day': client_day,
            'current_time': client_time
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def save_routine_slot(request):
    """Add or edit a single routine slot. Teachers or Admin can manage."""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            slot_id = data.get('id')

            dept = data.get('department', 'Computer Science & Technology')
            sem = data.get('semester', '5th Semester')
            shift = data.get('shift', '1st Shift')
            sec = data.get('section', 'A')
            day = data.get('day')
            start_time = data.get('start_time')
            end_time = data.get('end_time')
            subject = data.get('subject')
            subject_code = data.get('subject_code', '')
            teacher_initials = data.get('teacher_initials', '')
            teacher_name = data.get('teacher_name', '')
            room = data.get('room')
            email = data.get('email')

            if not day or not start_time or not end_time or not subject or not room:
                return JsonResponse({'status': 'error', 'message': 'Day, time, subject and room are required'}, status=400)

            # Optional teacher link
            teacher_obj = None
            if email:
                teacher_obj = UserAccount.objects.filter(email=email, role='teacher').first()

            time_slot = f"{start_time}-{end_time}"

            if slot_id:
                r = Routine.objects.get(id=slot_id)
                r.department = dept
                r.semester = sem
                r.shift = shift
                r.section = sec
                r.day = day
                r.start_time = start_time
                r.end_time = end_time
                r.time_slot = time_slot
                r.subject = subject
                r.subject_code = subject_code
                r.teacher_initials = teacher_initials
                r.teacher_name = teacher_name
                r.room = room
                if teacher_obj:
                    r.teacher = teacher_obj
                r.save()
                msg = 'Routine slot updated successfully!'
            else:
                r = Routine.objects.create(
                    department=dept,
                    semester=sem,
                    shift=shift,
                    section=sec,
                    day=day,
                    start_time=start_time,
                    end_time=end_time,
                    time_slot=time_slot,
                    subject=subject,
                    subject_code=subject_code,
                    teacher_initials=teacher_initials,
                    teacher_name=teacher_name,
                    room=room,
                    teacher=teacher_obj
                )
                msg = 'Routine slot added successfully!'

            return JsonResponse({'status': 'success', 'message': msg, 'id': r.id})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@csrf_exempt
def delete_routine_slot(request, slot_id):
    """Delete a routine slot."""
    if request.method == 'POST':
        try:
            Routine.objects.filter(id=slot_id).delete()
            return JsonResponse({'status': 'success', 'message': 'Routine slot deleted!'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@csrf_exempt
def upload_routine_file(request):
    """Upload PDF, Image (JPG/PNG), or Excel (.xlsx) file reference or bulk parse Excel into database."""
    if request.method == 'POST':
        try:
            title = request.POST.get('title', 'Class Routine')
            dept = request.POST.get('department', 'Computer Science & Technology')
            sem = request.POST.get('semester', '5th Semester')
            shift = request.POST.get('shift', '1st Shift')
            uploaded_file = request.FILES.get('file')

            if not uploaded_file:
                return JsonResponse({'status': 'error', 'message': 'No file uploaded'}, status=400)

            ext = uploaded_file.name.split('.')[-1].lower()
            file_type = 'excel' if ext in ['xlsx', 'xls'] else ('pdf' if ext == 'pdf' else 'image')

            rf = RoutineFile.objects.create(
                title=title,
                file_type=file_type,
                file=uploaded_file,
                department=dept,
                semester=sem,
                shift=shift
            )

            imported_count = 0
            parse_errors = []

            # If Excel, parse rows using openpyxl and import directly to Database
            if ext in ['xlsx', 'xls']:
                try:
                    import openpyxl
                    # Re-read file from saved path
                    wb = openpyxl.load_workbook(rf.file.path, data_only=True)
                    sheet = wb.active

                    # ---- SMART HEADER DETECTION ----
                    # Map common header names to our field keys
                    HEADER_MAP = {
                        'day': ['day', 'দিন', 'বার', 'days'],
                        'start_time': ['start_time', 'start time', 'start', 'from', 'time start', 'শুরু'],
                        'end_time': ['end_time', 'end time', 'end', 'to', 'time end', 'শেষ'],
                        'time_slot': ['time_slot', 'time slot', 'time', 'period', 'সময়'],
                        'subject_code': ['subject_code', 'subject code', 'code', 'sub code', 'কোড', 'বিষয় কোড'],
                        'subject': ['subject', 'subject name', 'sub', 'course', 'বিষয়', 'কোর্স'],
                        'teacher_initials': ['teacher_initials', 'teacher initials', 'initials', 'teacher', 'sir', 'instructor', 'শিক্ষক', 'teacher name'],
                        'room': ['room', 'room no', 'room number', 'lab', 'রুম', 'কক্ষ', 'venue'],
                        'section': ['section', 'sec', 'group', 'শাখা'],
                        'department': ['department', 'dept', 'বিভাগ'],
                        'semester': ['semester', 'sem', 'সেমিস্টার'],
                        'shift': ['shift', 'শিফট'],
                    }

                    def match_header(cell_value):
                        """Return our field key for a header cell value."""
                        if not cell_value:
                            return None
                        val = str(cell_value).strip().lower()
                        for key, aliases in HEADER_MAP.items():
                            for alias in aliases:
                                if val == alias or val.startswith(alias):
                                    return key
                        return None

                    # Find header row (first row with 2+ recognized headers)
                    header_row_idx = None
                    col_map = {}  # field_key -> column_index (0-based)
                    
                    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(10, sheet.max_row), values_only=False), start=1):
                        temp_map = {}
                        for col_idx, cell in enumerate(row):
                            key = match_header(cell.value)
                            if key and key not in temp_map:
                                temp_map[key] = col_idx
                        if len(temp_map) >= 2:  # At least 2 recognized headers
                            header_row_idx = row_idx
                            col_map = temp_map
                            break

                    if not header_row_idx:
                        # Fallback: assume standard order (Day, StartTime, EndTime, SubjectCode, Subject, TeacherInitials, Room)
                        header_row_idx = 1
                        fallback_keys = ['day', 'start_time', 'end_time', 'subject_code', 'subject', 'teacher_initials', 'room']
                        for i, key in enumerate(fallback_keys):
                            col_map[key] = i

                    print(f"[EXCEL IMPORT] Header row: {header_row_idx}, columns: {col_map}")

                    def safe_str(val):
                        if val is None:
                            return ''
                        return str(val).strip()

                    def parse_time(val):
                        """Parse various time formats to HH:MM."""
                        if not val:
                            return ''
                        s = str(val).strip()
                        # Handle datetime objects from Excel
                        import datetime
                        if isinstance(val, datetime.time):
                            return val.strftime('%H:%M')
                        if isinstance(val, datetime.datetime):
                            return val.strftime('%H:%M')
                        # Handle "8:00", "08:00", "8.00", "08.00 AM" etc.
                        s = s.replace('.', ':').replace(' AM', '').replace(' PM', '').replace(' am', '').replace(' pm', '')
                        parts = s.split(':')
                        if len(parts) == 2:
                            try:
                                h = int(parts[0])
                                m = int(parts[1])
                                return f"{h:02d}:{m:02d}"
                            except ValueError:
                                pass
                        return s

                    # Read data rows
                    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
                        if not row:
                            continue
                        # Skip completely empty rows
                        if all(c is None or str(c).strip() == '' for c in row):
                            continue

                        def get_col(key):
                            idx = col_map.get(key)
                            if idx is not None and idx < len(row):
                                return row[idx]
                            return None

                        r_day = safe_str(get_col('day')) or 'Sunday'
                        
                        # Handle time_slot column (e.g. "08:00-09:30")
                        r_st = ''
                        r_et = ''
                        time_slot_val = safe_str(get_col('time_slot'))
                        if time_slot_val and '-' in time_slot_val:
                            parts = time_slot_val.split('-')
                            r_st = parse_time(parts[0])
                            r_et = parse_time(parts[1])
                        else:
                            r_st = parse_time(get_col('start_time')) or '08:00'
                            r_et = parse_time(get_col('end_time')) or '08:45'

                        r_code = safe_str(get_col('subject_code'))
                        r_sub = safe_str(get_col('subject'))
                        r_teacher = safe_str(get_col('teacher_initials'))
                        r_room = safe_str(get_col('room')) or 'Room 304'
                        r_section = safe_str(get_col('section')) or 'A'
                        r_dept = safe_str(get_col('department')) or dept
                        r_sem = safe_str(get_col('semester')) or sem
                        r_shift = safe_str(get_col('shift')) or shift

                        # Skip rows without subject
                        if not r_sub:
                            continue

                        try:
                            Routine.objects.create(
                                department=r_dept,
                                semester=r_sem,
                                shift=r_shift,
                                section=r_section,
                                day=r_day,
                                start_time=r_st,
                                end_time=r_et,
                                time_slot=f"{r_st}-{r_et}",
                                subject_code=r_code,
                                subject=r_sub,
                                teacher_initials=r_teacher,
                                room=r_room
                            )
                            imported_count += 1
                        except Exception as row_err:
                            parse_errors.append(str(row_err))

                except Exception as ex_err:
                    print(f"[EXCEL IMPORT ERROR]: {ex_err}")
                    import traceback
                    traceback.print_exc()
                    parse_errors.append(str(ex_err))

            msg = f"Routine file uploaded! ({imported_count} slots imported from Excel)" if imported_count > 0 else "Routine file saved successfully!"
            if parse_errors:
                msg += f" ({len(parse_errors)} rows had errors)"
                
            return JsonResponse({
                'status': 'success',
                'message': msg,
                'file_url': rf.file.url,
                'file_type': rf.file_type,
                'imported_count': imported_count
            })
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@csrf_exempt
def get_routine_files(request):
    """Get list of reference routine files (PDF, Image, Excel)."""
    try:
        files = RoutineFile.objects.all().order_by('-uploaded_at')
        data = []
        for f in files:
            data.append({
                'id': f.id,
                'title': f.title,
                'file_type': f.file_type,
                'file_url': f.file.url,
                'department': f.department or '',
                'semester': f.semester or '',
                'shift': f.shift or '',
                'uploaded_at': f.uploaded_at.strftime('%Y-%m-%d %H:%M')
            })
        return JsonResponse({'status': 'success', 'data': data})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


# ==============================================================================
# ASSIGNMENT MANAGEMENT SYSTEM API
# ==============================================================================

@csrf_exempt
def create_assignment(request):
    """Teacher creates an assignment with file or Google Drive link."""
    if request.method == 'POST':
        try:
            email = request.POST.get('email')
            title = request.POST.get('title')
            description = request.POST.get('description', '')
            subject = request.POST.get('subject', 'General')
            subject_code = request.POST.get('subject_code', '')
            department = request.POST.get('department', 'Computer Science & Technology')
            semester = request.POST.get('semester', '5th Semester')
            shift = request.POST.get('shift', '1st Shift')
            total_marks = request.POST.get('total_marks', 100)
            due_date = request.POST.get('due_date', '')
            drive_link = request.POST.get('drive_link', '').strip() or None
            uploaded_file = request.FILES.get('file')

            if not title:
                return JsonResponse({'status': 'error', 'message': 'Assignment title is required'}, status=400)

            teacher = UserAccount.objects.filter(email=email, role='teacher').first()
            if not teacher:
                teacher = UserAccount.objects.filter(role='teacher').first()
            if not teacher:
                return JsonResponse({'status': 'error', 'message': 'No teacher account registered in system'}, status=403)

            assign = Assignment.objects.create(
                title=title,
                description=description,
                subject=subject,
                subject_code=subject_code,
                department=department,
                semester=semester,
                shift=shift,
                total_marks=total_marks,
                due_date=due_date,
                drive_link=drive_link,
                file=uploaded_file,
                posted_by=teacher
            )

            return JsonResponse({
                'status': 'success',
                'message': 'Assignment created & published to students successfully!',
                'id': assign.id
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


@csrf_exempt
def get_assignments(request):
    """Fetch assignments matching department, semester, shift filters with submission state if student email provided."""
    try:
        dept = request.GET.get('department')
        sem = request.GET.get('semester')
        shift = request.GET.get('shift')
        student_email = request.GET.get('student_email')
        teacher_email = request.GET.get('teacher_email')

        query = Assignment.objects.all()

        if dept and dept != 'All':
            dept_kw = 'Computer' if 'computer' in dept.lower() or 'cst' in dept.lower() else dept.split()[0]
            query = query.filter(models.Q(department__icontains=dept_kw) | models.Q(department='All') | models.Q(department=''))
        if sem and sem != 'All':
            query = query.filter(models.Q(semester__icontains=sem.split()[0]) | models.Q(semester='All') | models.Q(semester=''))
        if shift and shift != 'All':
            query = query.filter(models.Q(shift__icontains=shift.split()[0]) | models.Q(shift='All') | models.Q(shift=''))
        if teacher_email:
            query = query.filter(posted_by__email=teacher_email)

        assignments = query.order_by('-created_at')

        student = None
        if student_email:
            student = UserAccount.objects.filter(email=student_email, role='student').first()

        data = []
        for a in assignments:
            submission_info = None
            if student:
                sub = AssignmentSubmission.objects.filter(assignment=a, student=student).first()
                if sub:
                    submission_info = {
                        'id': sub.id,
                        'file_url': sub.submission_file.url if sub.submission_file else None,
                        'drive_link': sub.drive_link or '',
                        'notes': sub.notes or '',
                        'submitted_at': sub.submitted_at.strftime('%d %b, %Y %I:%M %p'),
                        'marks_obtained': sub.marks_obtained or '--',
                        'feedback': sub.feedback or '',
                        'status': sub.status
                    }

            total_submissions = a.submissions.count()

            data.append({
                'id': a.id,
                'title': a.title,
                'description': a.description,
                'subject': a.subject,
                'subject_code': a.subject_code or '',
                'department': a.department,
                'semester': a.semester,
                'shift': a.shift,
                'total_marks': a.total_marks,
                'due_date': a.due_date,
                'file_url': a.file.url if a.file else None,
                'drive_link': a.drive_link or '',
                'posted_by': f"{a.posted_by.first_name} {a.posted_by.last_name}",
                'created_at': a.created_at.strftime('%d %b, %Y'),
                'submission': submission_info,
                'is_submitted': submission_info is not None,
                'total_submissions': total_submissions
            })

        return JsonResponse({'status': 'success', 'data': data})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@csrf_exempt
def delete_assignment(request, assignment_id):
    """Delete an assignment."""
    if request.method == 'POST':
        try:
            Assignment.objects.filter(id=assignment_id).delete()
            return JsonResponse({'status': 'success', 'message': 'Assignment deleted successfully!'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


@csrf_exempt
def submit_assignment(request):
    """Student submits completed assignment via File or Google Drive Link."""
    if request.method == 'POST':
        try:
            assignment_id = request.POST.get('assignment_id')
            email = request.POST.get('email')
            drive_link = request.POST.get('drive_link', '').strip() or None
            notes = request.POST.get('notes', '')
            submission_file = request.FILES.get('file')

            if not assignment_id or not email:
                return JsonResponse({'status': 'error', 'message': 'Missing assignment ID or student email'}, status=400)

            student = UserAccount.objects.filter(email=email, role='student').first()
            if not student:
                student = UserAccount.objects.filter(role='student').first()
            if not student:
                return JsonResponse({'status': 'error', 'message': 'Student account not found'}, status=404)

            assignment = Assignment.objects.get(id=assignment_id)

            sub, created = AssignmentSubmission.objects.update_or_create(
                assignment=assignment,
                student=student,
                defaults={
                    'drive_link': drive_link,
                    'notes': notes,
                    'status': 'Submitted'
                }
            )
            if submission_file:
                sub.submission_file = submission_file
                sub.save()

            return JsonResponse({'status': 'success', 'message': 'Assignment submitted successfully!'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


@csrf_exempt
def get_assignment_submissions(request, assignment_id):
    """Teacher views all student submissions for an assignment."""
    try:
        assignment = Assignment.objects.get(id=assignment_id)
        submissions = assignment.submissions.all().order_by('-submitted_at')

        sub_list = []
        for s in submissions:
            sub_list.append({
                'id': s.id,
                'student_id': s.student.id,
                'student_name': f"{s.student.first_name} {s.student.last_name}",
                'roll': s.student.roll or '--',
                'file_url': s.submission_file.url if s.submission_file else None,
                'drive_link': s.drive_link or '',
                'notes': s.notes or '',
                'submitted_at': s.submitted_at.strftime('%d %b, %Y %I:%M %p'),
                'marks_obtained': s.marks_obtained or '',
                'feedback': s.feedback or '',
                'status': s.status
            })

        return JsonResponse({
            'status': 'success',
            'assignment_title': assignment.title,
            'total_marks': assignment.total_marks,
            'data': sub_list
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@csrf_exempt
def grade_submission(request):
    """Teacher grades a student assignment submission."""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            submission_id = data.get('submission_id')
            marks = data.get('marks_obtained')
            feedback = data.get('feedback', '')

            sub = AssignmentSubmission.objects.get(id=submission_id)
            sub.marks_obtained = str(marks)
            sub.feedback = feedback
            sub.status = 'Graded'
            sub.save()

            return JsonResponse({'status': 'success', 'message': 'Submission graded successfully!'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


# ==============================================================================
# PUBLIC INFO API (No login required)
# ==============================================================================

def get_public_college_info(request):
    """Public endpoint — college stats without login."""
    try:
        total_teachers = UserAccount.objects.filter(role='teacher').count()
        total_students = UserAccount.objects.filter(role='student').count()
        return JsonResponse({
            'status': 'success',
            'data': {
                'total_teachers': total_teachers,
                'total_students': total_students,
            }
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


def get_public_teachers(request):
    """Public teacher list — no login needed."""
    try:
        teachers = UserAccount.objects.filter(role='teacher').order_by('first_name')
        teacher_list = []
        for t in teachers:
            teacher_list.append({
                'id': t.id,
                'name': f"{t.first_name} {t.last_name}",
                'department': t.department or 'Computer Science & Technology',
                'designation': t.designation or 'Instructor',
                'qualification': t.qualification or '',
                'specialized_subjects': t.specialized_subjects or '',
                'room_number': t.room_number or '',
                'profile_picture': t.profile_picture.url if t.profile_picture else None,
            })
        return JsonResponse({'status': 'success', 'data': teacher_list})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


# ==============================================================================
# CR (CLASS REPRESENTATIVE) SYSTEM
# ==============================================================================

def get_crs(request):
    """Get active/approved CR list — with optional filters."""
    try:
        semester = request.GET.get('semester')
        group = request.GET.get('group')
        gender = request.GET.get('gender')
        academic_year = request.GET.get('academic_year')
        include_pending = request.GET.get('include_pending') == 'true'

        query = ClassRepresentative.objects.all()
        if not include_pending:
            query = query.filter(is_active=True, is_approved=True)

        if semester and semester != 'All':
            query = query.filter(semester__icontains=semester.split()[0])
        if group and group != 'All':
            query = query.filter(group__iexact=group)
        if gender and gender != 'All':
            query = query.filter(gender__iexact=gender)
        if academic_year and academic_year != 'All':
            query = query.filter(academic_year=academic_year)

        cr_list = []
        for cr in query.order_by('-is_approved', 'semester', 'group', 'gender'):
            cr_list.append({
                'id': cr.id,
                'student_id': cr.student.id,
                'name': f"{cr.student.first_name} {cr.student.last_name}",
                'roll': cr.student.roll or '--',
                'mobile': cr.student.mobile or '',
                'email': cr.student.email or '',
                'semester': cr.semester,
                'group': cr.group or '',
                'gender': cr.gender,
                'batch': cr.batch or cr.student.batch or '',
                'academic_year': cr.academic_year,
                'is_active': cr.is_active,
                'is_approved': cr.is_approved,
                'profile_picture': cr.student.profile_picture.url if cr.student.profile_picture else None,
            })
        return JsonResponse({'status': 'success', 'data': cr_list})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@csrf_exempt
def nominate_cr(request):
    """Student nominates themselves for CR (requires Teacher Approval)."""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            student_email = data.get('student_email') or data.get('email')
            roll = data.get('roll')
            semester = data.get('semester', '1st Semester')
            group = data.get('group', '')
            gender = data.get('gender', 'Boys')

            student = None
            if student_email:
                student = UserAccount.objects.filter(email=student_email).first()
            if not student and roll:
                student = UserAccount.objects.filter(roll=roll).first()

            if not student:
                return JsonResponse({'status': 'error', 'message': 'Student account not found. Please ensure you are logged in.'}, status=404)

            # Check if active or pending nomination exists
            existing = ClassRepresentative.objects.filter(student=student, is_active=True).first()
            if existing:
                if existing.is_approved:
                    return JsonResponse({'status': 'error', 'message': 'You are already an approved CR!'}, status=400)
                else:
                    return JsonResponse({'status': 'error', 'message': 'Your CR nomination is pending teacher approval.'}, status=400)

            cr = ClassRepresentative.objects.create(
                student=student, semester=semester, group=group or '', gender=gender,
                department=student.department or 'Computer Science & Technology',
                batch=student.batch or '', is_active=True, is_approved=False
            )
            return JsonResponse({'status': 'success', 'message': 'CR Nomination submitted! Awaiting Teacher Approval.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)


@csrf_exempt
def approve_cr(request, cr_id):
    """Teacher approves a pending CR nomination."""
    if request.method == 'POST':
        try:
            cr = ClassRepresentative.objects.get(id=cr_id)
            # Deactivate previous approved CR for same sem/group/gender
            ClassRepresentative.objects.filter(
                semester=cr.semester, group=cr.group, gender=cr.gender,
                is_active=True, is_approved=True
            ).exclude(id=cr.id).update(is_active=False)

            cr.is_approved = True
            cr.is_active = True
            cr.save()
            return JsonResponse({'status': 'success', 'message': f'CR nomination for {cr.student.first_name} approved!'})
        except ClassRepresentative.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'CR nomination not found'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)


@csrf_exempt
def assign_cr(request):
    """Teacher directly assigns and approves a student as CR."""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            teacher_email = data.get('teacher_email')
            student_id = data.get('student_id')
            semester = data.get('semester')
            group = data.get('group', '')
            gender = data.get('gender')
            batch = data.get('batch', '')
            academic_year = data.get('academic_year', '2025-2026')

            teacher = UserAccount.objects.filter(email=teacher_email, role='teacher').first()
            if not teacher:
                return JsonResponse({'status': 'error', 'message': 'Unauthorized.'}, status=403)

            student = UserAccount.objects.get(id=student_id, role='student')

            if not semester:
                semester = student.semester or '1st Semester'

            ClassRepresentative.objects.filter(
                semester=semester, group=group, gender=gender,
                academic_year=academic_year, is_active=True
            ).update(is_active=False)

            cr = ClassRepresentative.objects.create(
                student=student, semester=semester, group=group, gender=gender,
                department=student.department or 'Computer Science & Technology',
                batch=batch or student.batch or '', academic_year=academic_year,
                is_active=True, is_approved=True, assigned_by=teacher
            )
            return JsonResponse({'status': 'success', 'message': f'{student.first_name} assigned & approved as CR!', 'id': cr.id})
        except UserAccount.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Student not found'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


@csrf_exempt
def remove_cr(request, cr_id):
    """Teacher removes/deactivates a CR."""
    if request.method == 'POST':
        try:
            cr = ClassRepresentative.objects.get(id=cr_id)
            cr.is_active = False
            cr.save()
            return JsonResponse({'status': 'success', 'message': 'CR removed!'})
        except ClassRepresentative.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'CR not found'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


def get_leaderboard(request):
    """Get top performing students based on Quiz submissions & Attendance."""
    try:
        dept = request.GET.get('department')
        students = UserAccount.objects.filter(role='student')
        if dept and dept != 'All':
            students = students.filter(department__icontains=dept.split()[0])

        leaderboard = []
        for s in students:
            # Submissions total score
            submissions = QuizSubmission.objects.filter(student=s)
            quiz_count = submissions.count()
            avg_quiz_score = int(sum(sub.score for sub in submissions) / quiz_count) if quiz_count else 0
            
            # Attendance
            total_att = Attendance.objects.filter(student=s).count()
            present_att = Attendance.objects.filter(student=s, status='P').count()
            att_rate = int((present_att / total_att) * 100) if total_att else 0

            # Combined score (when no data at all, total=0)
            total_score = round((avg_quiz_score * 0.7) + (att_rate * 0.3), 1)

            leaderboard.append({
                'id': s.id,
                'name': f"{s.first_name} {s.last_name}",
                'roll': s.roll or '--',
                'department': s.department or 'CST',
                'semester': s.semester or '1st Semester',
                'quiz_avg': avg_quiz_score,
                'quiz_count': quiz_count,
                'attendance_rate': att_rate,
                'overall_score': total_score,
                'total_score': total_score,
                'profile_picture': s.profile_picture.url if s.profile_picture else None,
            })

        # Sort by overall_score descending
        leaderboard.sort(key=lambda x: x['overall_score'], reverse=True)
        # Assign ranks
        for rank, entry in enumerate(leaderboard, 1):
            entry['rank'] = rank

        return JsonResponse({'status': 'success', 'data': leaderboard[:15]})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


# ==============================================================================
# SEMESTER BOOKS API
# ==============================================================================

def get_semester_books(request):
    """Get books list — filtered by semester."""
    try:
        semester = request.GET.get('semester')
        department = request.GET.get('department')
        query = SemesterBook.objects.all()
        if semester and semester != 'All':
            query = query.filter(semester__icontains=semester.split()[0])
        if department and department != 'All':
            query = query.filter(department__icontains=department.split()[0])
        books = []
        for b in query.order_by('semester', 'subject_name'):
            books.append({
                'id': b.id,
                'semester': b.semester,
                'subject_name': b.subject_name,
                'subject_code': b.subject_code or '',
                'author': b.author or '',
                'publisher': b.publisher or '',
                'notes': b.notes or '',
                'shift': b.shift or '',
                'added_by': f"{b.added_by.first_name} {b.added_by.last_name}" if b.added_by else 'Admin',
            })
        return JsonResponse({'status': 'success', 'data': books})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@csrf_exempt
def add_semester_book(request):
    """Teacher adds a book to a semester."""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            teacher_email = data.get('teacher_email')
            teacher = UserAccount.objects.filter(email=teacher_email, role='teacher').first()
            if not teacher:
                return JsonResponse({'status': 'error', 'message': 'Unauthorized'}, status=403)
            book = SemesterBook.objects.create(
                semester=data.get('semester'), subject_name=data.get('subject_name'),
                subject_code=data.get('subject_code', ''), author=data.get('author', ''),
                publisher=data.get('publisher', ''), notes=data.get('notes', ''),
                shift=data.get('shift', ''),
                department=data.get('department', 'Computer Science & Technology'),
                added_by=teacher
            )
            return JsonResponse({'status': 'success', 'message': 'Book added!', 'id': book.id})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


@csrf_exempt
def delete_semester_book(request, book_id):
    """Teacher deletes a book entry."""
    if request.method == 'POST':
        try:
            SemesterBook.objects.filter(id=book_id).delete()
            return JsonResponse({'status': 'success', 'message': 'Book deleted!'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


# ==============================================================================
# MESSAGE SYSTEM API
# ==============================================================================

@csrf_exempt
def send_message(request):
    """Student sends a message to a teacher."""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            sender_email = data.get('sender_email')
            receiver_id = data.get('receiver_id')
            subject = data.get('subject', '')
            content = data.get('content')
            parent_id = data.get('parent_id')
            sender = UserAccount.objects.get(email=sender_email)
            receiver = UserAccount.objects.get(id=receiver_id)
            parent = Message.objects.filter(id=parent_id).first() if parent_id else None
            msg = Message.objects.create(
                sender=sender, receiver=receiver, subject=subject,
                content=content, parent_message=parent
            )
            return JsonResponse({'status': 'success', 'message': 'Message sent!', 'id': msg.id})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


def get_messages(request):
    """Get inbox/sent messages for a user."""
    try:
        email = request.GET.get('email')
        box = request.GET.get('box', 'inbox')
        user = UserAccount.objects.get(email=email)
        if box == 'inbox':
            msgs = Message.objects.filter(receiver=user, parent_message__isnull=True).order_by('-sent_at')
        else:
            msgs = Message.objects.filter(sender=user, parent_message__isnull=True).order_by('-sent_at')
        data = []
        for m in msgs:
            replies = Message.objects.filter(parent_message=m).order_by('sent_at')
            reply_list = [{'id': r.id, 'sender_name': f"{r.sender.first_name} {r.sender.last_name}",
                           'sender_role': r.sender.role, 'content': r.content,
                           'sent_at': r.sent_at.strftime('%d %b %Y, %I:%M %p')} for r in replies]
            data.append({
                'id': m.id, 'sender_id': m.sender.id,
                'sender_name': f"{m.sender.first_name} {m.sender.last_name}",
                'sender_role': m.sender.role,
                'sender_pic': m.sender.profile_picture.url if m.sender.profile_picture else None,
                'receiver_id': m.receiver.id,
                'receiver_name': f"{m.receiver.first_name} {m.receiver.last_name}",
                'receiver_role': m.receiver.role,
                'subject': m.subject or '(No subject)',
                'content': m.content,
                'sent_at': m.sent_at.strftime('%d %b %Y, %I:%M %p'),
                'is_read': m.is_read, 'replies': reply_list, 'reply_count': len(reply_list),
            })
        if box == 'inbox':
            Message.objects.filter(receiver=user, is_read=False).update(is_read=True)
        unread_count = Message.objects.filter(receiver=user, is_read=False).count()
        return JsonResponse({'status': 'success', 'data': data, 'unread_count': unread_count})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@csrf_exempt
def reply_message(request):
    """Reply to a message."""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            sender_email = data.get('sender_email')
            parent_id = data.get('parent_id')
            content = data.get('content')
            parent = Message.objects.get(id=parent_id)
            sender = UserAccount.objects.get(email=sender_email)
            receiver = parent.sender if parent.receiver.email == sender_email else parent.receiver
            msg = Message.objects.create(
                sender=sender, receiver=receiver,
                subject=f"Re: {parent.subject or '(No subject)'}",
                content=content, parent_message=parent
            )
            return JsonResponse({'status': 'success', 'message': 'Reply sent!', 'id': msg.id})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


def get_unread_count(request):
    """Get unread message count."""
    try:
        email = request.GET.get('email')
        user = UserAccount.objects.get(email=email)
        count = Message.objects.filter(receiver=user, is_read=False).count()
        return JsonResponse({'status': 'success', 'unread_count': count})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


# ==============================================================================
# COMPLAINT BOX API
# ==============================================================================

@csrf_exempt
def submit_complaint(request):
    """Student submits an anonymous complaint."""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            student_email = data.get('student_email')
            content = data.get('content', '').strip()
            category = data.get('category', 'General')
            if not content:
                return JsonResponse({'status': 'error', 'message': 'Content required'}, status=400)
            student = UserAccount.objects.get(email=student_email, role='student')
            complaint = Complaint.objects.create(student=student, content=content, category=category)
            return JsonResponse({'status': 'success', 'message': 'Complaint submitted anonymously!', 'id': complaint.id})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


def get_complaints(request):
    """Get complaint list. Teachers see full info; others see anonymous view."""
    try:
        viewer_email = request.GET.get('email')
        is_teacher = False
        if viewer_email:
            viewer = UserAccount.objects.filter(email=viewer_email).first()
            if viewer and viewer.role == 'teacher':
                is_teacher = True
        complaints = Complaint.objects.all().order_by('-submitted_at')
        data = []
        for c in complaints:
            item = {
                'id': c.id, 'content': c.content, 'category': c.category or 'General',
                'submitted_at': c.submitted_at.strftime('%d %b %Y, %I:%M %p'),
                'is_resolved': c.is_resolved, 'response': c.response or '',
                'responded_by': f"{c.responded_by.first_name} {c.responded_by.last_name}" if c.responded_by else '',
                'responded_at': c.responded_at.strftime('%d %b %Y') if c.responded_at else '',
            }
            if is_teacher:
                item['student_name'] = f"{c.student.first_name} {c.student.last_name}"
                item['student_roll'] = c.student.roll or '--'
                item['student_email'] = c.student.email
            else:
                item['student_name'] = 'Anonymous'
                item['student_roll'] = '***'
            data.append(item)
        return JsonResponse({'status': 'success', 'data': data})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@csrf_exempt
def respond_complaint(request, complaint_id):
    """Teacher responds to a complaint."""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            teacher_email = data.get('teacher_email')
            response_text = data.get('response')
            teacher = UserAccount.objects.filter(email=teacher_email, role='teacher').first()
            if not teacher:
                return JsonResponse({'status': 'error', 'message': 'Unauthorized'}, status=403)
            complaint = Complaint.objects.get(id=complaint_id)
            complaint.response = response_text
            complaint.responded_by = teacher
            complaint.responded_at = timezone.now()
            complaint.is_resolved = True
            complaint.save()
            return JsonResponse({'status': 'success', 'message': 'Response added!'})
        except Complaint.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Complaint not found'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


# ==============================================================================
# QUIZ SYSTEM API
# ==============================================================================

@csrf_exempt
def create_quiz(request):
    """Teacher creates a quiz with questions."""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            teacher_email = data.get('teacher_email')
            teacher = UserAccount.objects.filter(email=teacher_email, role='teacher').first()
            if not teacher:
                teacher = UserAccount.objects.filter(role='teacher').first()
            if not teacher:
                return JsonResponse({'status': 'error', 'message': 'Unauthorized'}, status=403)

            now = timezone.now()
            start_raw = data.get('start_time')
            end_raw = data.get('end_time')

            start_time = now - timezone.timedelta(minutes=5)
            end_time = now + timezone.timedelta(days=30)

            if start_raw:
                try:
                    start_time = timezone.datetime.fromisoformat(str(start_raw))
                    if timezone.is_naive(start_time):
                        start_time = timezone.make_aware(start_time)
                except Exception:
                    pass

            if end_raw:
                try:
                    end_time = timezone.datetime.fromisoformat(str(end_raw))
                    if timezone.is_naive(end_time):
                        end_time = timezone.make_aware(end_time)
                except Exception:
                    pass

            quiz = Quiz.objects.create(
                title=data.get('title', 'Class Quiz'), description=data.get('description', ''),
                subject=data.get('subject', 'General'),
                department=data.get('department', 'Computer Science & Technology'),
                semester=data.get('semester', '5th Semester'), language=data.get('language', 'English'),
                duration_minutes=int(data.get('duration_minutes', 30)),
                total_marks=int(data.get('total_marks', 100)),
                start_time=start_time, end_time=end_time,
                is_active=True, created_by=teacher
            )
            for i, q in enumerate(data.get('questions', [])):
                QuizQuestion.objects.create(
                    quiz=quiz, question_text=q.get('question_text'),
                    option_a=q.get('option_a'), option_b=q.get('option_b'),
                    option_c=q.get('option_c'), option_d=q.get('option_d'),
                    correct_answer=q.get('correct_answer', 'A').upper(),
                    marks=int(q.get('marks', 5)), order=i + 1
                )
            return JsonResponse({'status': 'success', 'message': f'Quiz created with {len(data.get("questions",[]))} questions!', 'id': quiz.id})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


def get_quizzes(request):
    """Get quizzes list."""
    try:
        semester = request.GET.get('semester')
        email = request.GET.get('email')
        teacher_email = request.GET.get('teacher_email')
        now = timezone.now()
        query = Quiz.objects.filter(is_active=True)
        if semester and semester != 'All':
            query = query.filter(semester__icontains=semester.split()[0])
        if teacher_email:
            query = query.filter(created_by__email=teacher_email)
        student = UserAccount.objects.filter(email__iexact=email).first() if email else None
        data = []
        for q in query.order_by('-created_at'):
            # Determine status — treat missing times as open
            st = q.start_time
            et = q.end_time
            if st and et:
                if now < st:
                    status_label = 'upcoming'
                elif now > et:
                    status_label = 'ended'
                else:
                    status_label = 'active'
            elif st and not et:
                status_label = 'active' if now >= st else 'upcoming'
            elif et and not st:
                status_label = 'active' if now <= et else 'ended'
            else:
                status_label = 'active'  # No time limits — always active
            sub_info = None
            if student:
                sub = QuizSubmission.objects.filter(quiz=q, student=student).first()
                if sub:
                    sub_info = {'id': sub.id, 'score': sub.score, 'correct_count': sub.correct_count,
                                'total_questions': sub.total_questions, 'is_completed': sub.is_completed,
                                'time_taken': sub.time_taken_minutes,
                                'submitted_at': sub.submitted_at.strftime('%d %b %Y, %I:%M %p') if sub.submitted_at else None}
            data.append({
                'id': q.id, 'title': q.title, 'description': q.description or '',
                'subject': q.subject, 'semester': q.semester, 'language': q.language,
                'duration_minutes': q.duration_minutes, 'total_marks': q.total_marks,
                'questions_count': q.questions.count(),
                'start_time': st.strftime('%d %b %Y, %I:%M %p') if st else 'Open',
                'end_time': et.strftime('%d %b %Y, %I:%M %p') if et else 'Open',
                'start_time_iso': st.isoformat() if st else None,
                'end_time_iso': et.isoformat() if et else None,
                'status': status_label,
                'created_by': f"{q.created_by.first_name} {q.created_by.last_name}",
                'submission': sub_info, 'is_submitted': sub_info is not None and sub_info['is_completed'],
            })
        return JsonResponse({'status': 'success', 'data': data})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


def get_quiz_questions(request, quiz_id):
    """Student gets quiz questions (without correct answers)."""
    try:
        email = request.GET.get('email')
        student = UserAccount.objects.filter(email=email, role='student').first()
        if not student:
            return JsonResponse({'status': 'error', 'message': 'Unauthorized'}, status=403)
        quiz = Quiz.objects.get(id=quiz_id, is_active=True)
        now = timezone.now()
        if now < quiz.start_time:
            return JsonResponse({'status': 'error', 'message': 'Quiz has not started yet!'}, status=400)
        if now > quiz.end_time:
            return JsonResponse({'status': 'error', 'message': 'Quiz has ended!'}, status=400)
        existing = QuizSubmission.objects.filter(quiz=quiz, student=student, is_completed=True).first()
        if existing:
            return JsonResponse({'status': 'error', 'message': 'You have already submitted this quiz!'}, status=400)
        sub, created = QuizSubmission.objects.get_or_create(
            quiz=quiz, student=student, defaults={'total_questions': quiz.questions.count()}
        )
        questions = quiz.questions.all()
        q_list = [{'id': q.id, 'question_text': q.question_text, 'option_a': q.option_a,
                   'option_b': q.option_b, 'option_c': q.option_c, 'option_d': q.option_d,
                   'marks': q.marks, 'order': q.order} for q in questions]
        end_by_duration = sub.started_at + timezone.timedelta(minutes=quiz.duration_minutes)
        actual_end = min(quiz.end_time, end_by_duration)
        remaining_seconds = max(0, int((actual_end - now).total_seconds()))
        return JsonResponse({
            'status': 'success',
            'quiz': {'id': quiz.id, 'title': quiz.title, 'subject': quiz.subject,
                     'language': quiz.language, 'duration_minutes': quiz.duration_minutes,
                     'total_marks': quiz.total_marks},
            'questions': q_list, 'remaining_seconds': remaining_seconds,
            'submission_id': sub.id, 'existing_answers': sub.answers,
        })
    except Quiz.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Quiz not found'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@csrf_exempt
def start_quiz(request):
    """Alias: Student starts a quiz — returns quiz info + questions (no correct answers)."""
    quiz_id = request.GET.get('quiz_id')
    email = request.GET.get('email')
    try:
        student = None
        if email:
            student = UserAccount.objects.filter(email__iexact=email).first()
        if not student:
            student = UserAccount.objects.filter(role='student').first() or UserAccount.objects.first()
        
        quiz = Quiz.objects.get(id=quiz_id)
        now = timezone.now()
        
        if student:
            existing = QuizSubmission.objects.filter(quiz=quiz, student=student, is_completed=True).first()
            if existing:
                return JsonResponse({'status': 'error', 'message': 'You have already submitted this quiz!'}, status=400)
            sub, _ = QuizSubmission.objects.get_or_create(
                quiz=quiz, student=student, defaults={'total_questions': quiz.questions.count()}
            )
        else:
            sub = None

        questions = quiz.questions.all().order_by('order', 'id')
        q_list = [{
            'id': q.id, 'question_text': q.question_text,
            'option_a': q.option_a, 'option_b': q.option_b,
            'option_c': q.option_c, 'option_d': q.option_d,
            'marks': q.marks
        } for q in questions]

        # Calculate remaining seconds
        remaining_seconds = (quiz.duration_minutes or 30) * 60
        if quiz.end_time and sub:
            end_by_duration = sub.started_at + timezone.timedelta(minutes=quiz.duration_minutes or 30)
            actual_end = min(quiz.end_time, end_by_duration)
            remaining_seconds = max(0, int((actual_end - now).total_seconds()))

        return JsonResponse({
            'status': 'success',
            'data': {
                'id': quiz.id, 'title': quiz.title, 'subject': quiz.subject,
                'semester': quiz.semester, 'duration_minutes': quiz.duration_minutes,
                'total_marks': quiz.total_marks,
                'questions': q_list,
                'remaining_seconds': remaining_seconds
            }
        })
    except Quiz.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Quiz not found'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@csrf_exempt
def submit_quiz(request):
    """Student submits quiz answers."""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email = data.get('email') or data.get('student_email')
            quiz_id = data.get('quiz_id')
            raw_answers = data.get('answers', {})

            # Normalize answers to dict {q_id_str: option}
            answers = {}
            if isinstance(raw_answers, list):
                for item in raw_answers:
                    if isinstance(item, dict) and 'question_id' in item:
                        answers[str(item['question_id'])] = item.get('selected_option', '')
            elif isinstance(raw_answers, dict):
                answers = {str(k): str(v) for k, v in raw_answers.items()}

            student = None
            if email:
                student = UserAccount.objects.filter(email__iexact=email).first()
            if not student:
                student = UserAccount.objects.filter(role='student').first() or UserAccount.objects.first()

            quiz = Quiz.objects.get(id=quiz_id)
            sub, _ = QuizSubmission.objects.get_or_create(
                quiz=quiz, student=student, defaults={'total_questions': quiz.questions.count()}
            )
            if sub.is_completed:
                return JsonResponse({'status': 'error', 'message': 'Quiz already submitted!'}, status=400)
            
            questions = quiz.questions.all()
            correct_count = 0
            score = 0
            for q in questions:
                chosen = str(answers.get(str(q.id), '')).upper()
                if chosen and chosen == str(q.correct_answer).upper():
                    correct_count += 1
                    score += q.marks
            now = timezone.now()
            time_taken = (now - sub.started_at).total_seconds() / 60 if sub.started_at else 0
            sub.answers = answers
            sub.score = score
            sub.correct_count = correct_count
            sub.total_questions = questions.count()
            sub.submitted_at = now
            sub.time_taken_minutes = round(time_taken, 2)
            sub.is_completed = True
            sub.save()
            return JsonResponse({
                'status': 'success', 'message': 'Quiz submitted!',
                'score': score, 'correct_count': correct_count,
                'total_questions': questions.count(), 'total_marks': quiz.total_marks,
                'percentage': round((score / quiz.total_marks) * 100, 1) if quiz.total_marks > 0 else 0,
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


def get_quiz_results(request, quiz_id):
    """Teacher sees all student results for a quiz."""
    try:
        quiz = Quiz.objects.get(id=quiz_id)
        submissions = QuizSubmission.objects.filter(quiz=quiz, is_completed=True).select_related('student')
        results = [{'student_name': f"{s.student.first_name} {s.student.last_name}",
                    'roll': s.student.roll or '--', 'score': s.score,
                    'correct_count': s.correct_count, 'total_questions': s.total_questions,
                    'time_taken': s.time_taken_minutes,
                    'submitted_at': s.submitted_at.strftime('%d %b %Y, %I:%M %p') if s.submitted_at else '',
                    'percentage': round((s.score / quiz.total_marks) * 100, 1) if quiz.total_marks > 0 else 0,
                    } for s in submissions.order_by('-score')]
        return JsonResponse({'status': 'success', 'quiz_title': quiz.title,
                             'total_marks': quiz.total_marks, 'total_submissions': len(results), 'data': results})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@csrf_exempt
def delete_quiz(request, quiz_id):
    """Teacher deletes a quiz."""
    if request.method == 'POST':
        try:
            Quiz.objects.filter(id=quiz_id).delete()
            return JsonResponse({'status': 'success', 'message': 'Quiz deleted!'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


@csrf_exempt
def save_quiz_progress(request):
    """Auto-save student's quiz answers while in progress."""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email = data.get('email')
            quiz_id = data.get('quiz_id')
            answers = data.get('answers', {})
            student = UserAccount.objects.get(email=email, role='student')
            quiz = Quiz.objects.get(id=quiz_id)
            sub, _ = QuizSubmission.objects.get_or_create(
                quiz=quiz, student=student, defaults={'total_questions': quiz.questions.count()}
            )
            if not sub.is_completed:
                sub.answers = answers
                sub.save()
            return JsonResponse({'status': 'success', 'message': 'Progress saved!'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


@csrf_exempt
def upload_routine_file(request):
    """Upload routine image or PDF file."""
    if request.method == 'POST':
        try:
            title = request.POST.get('title', 'Class Routine Image')
            department = request.POST.get('department', 'Computer Science & Technology')
            semester = request.POST.get('semester', '5th Semester')
            shift = request.POST.get('shift', '1st Shift')
            routine_file = request.FILES.get('file')

            if not routine_file:
                return JsonResponse({'status': 'error', 'message': 'No file uploaded!'}, status=400)

            ext = routine_file.name.split('.')[-1].lower()
            file_type = 'image' if ext in ['jpg', 'jpeg', 'png', 'webp', 'gif'] else ('pdf' if ext == 'pdf' else 'file')

            obj = RoutineFile.objects.create(
                title=title,
                file_type=file_type,
                file=routine_file,
                department=department,
                semester=semester,
                shift=shift
            )
            return JsonResponse({
                'status': 'success',
                'message': 'Routine image/file uploaded successfully!',
                'file_url': obj.file.url,
                'file_type': obj.file_type
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


@csrf_exempt
def get_routine_files(request):
    """Get uploaded routine images/PDFs for student's department & semester."""
    if request.method == 'GET':
        try:
            department = request.GET.get('department', '')
            semester = request.GET.get('semester', '')
            query = RoutineFile.objects.all().order_by('-uploaded_at')

            if department and department != 'All':
                query = query.filter(models.Q(department__icontains=department) | models.Q(department__isnull=True) | models.Q(department=''))
            if semester and semester != 'All':
                query = query.filter(models.Q(semester__icontains=semester) | models.Q(semester__isnull=True) | models.Q(semester=''))

            files_list = [{
                'id': f.id,
                'title': f.title,
                'file_type': f.file_type,
                'file_url': f.file.url if f.file else '',
                'department': f.department,
                'semester': f.semester,
                'shift': f.shift,
                'uploaded_at': f.uploaded_at.strftime('%d %b %Y, %I:%M %p') if f.uploaded_at else ''
            } for f in query]

            return JsonResponse({'status': 'success', 'data': files_list})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

